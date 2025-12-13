"""
Report generation tool for time tracking.

Generates timesheet reports with billable hours analysis.
Supports week-to-date, month-to-date, and custom date ranges.
"""

import os
from typing import Optional
from datetime import datetime, timedelta
from pathlib import Path

from gcalendar_mcp.tools.time_tracking.database import (
    ensure_database,
    get_setting,
    get_norm,
    list_projects,
)
from gcalendar_mcp.tools.time_tracking.parser import parse_events_batch, TimeEntry
from gcalendar_mcp.api.client import get_service


REPORTS_DIR = Path.home() / "Downloads" / "Timesheet reports"


def get_reports_dir() -> Path:
    """Get reports directory, creating if needed."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return REPORTS_DIR


def _get_period_bounds(
    period_type: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> tuple[datetime, datetime, int, int]:
    """
    Calculate period boundaries and workday counts.
    
    Returns: (start_datetime, end_datetime, workdays_in_period, workdays_elapsed)
    """
    now = datetime.now()
    today = now.date()
    
    if period_type == "week":
        # Monday of current week
        start = now - timedelta(days=now.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Count workdays
        workdays_total = 5  # Full week
        workdays_elapsed = 0
        iter_date = start.date()
        while iter_date <= today:
            if iter_date.weekday() < 5:
                workdays_elapsed += 1
            iter_date += timedelta(days=1)
        
    elif period_type == "month":
        # First day of current month
        start = datetime(now.year, now.month, 1)
        end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Count workdays in full month
        workdays_total = 0
        workdays_elapsed = 0
        iter_date = start.date()
        
        # Get last day of month
        if now.month == 12:
            last_day = datetime(now.year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(now.year, now.month + 1, 1) - timedelta(days=1)
        
        while iter_date <= last_day.date():
            if iter_date.weekday() < 5:
                workdays_total += 1
                if iter_date <= today:
                    workdays_elapsed += 1
            iter_date += timedelta(days=1)
        
    elif period_type == "custom":
        if not start_date or not end_date:
            raise ValueError("start_date and end_date required for custom period")
        
        start = datetime.fromisoformat(start_date)
        end = datetime.fromisoformat(end_date)
        if end.hour == 0 and end.minute == 0:
            end = end.replace(hour=23, minute=59, second=59)
        
        # Count workdays
        workdays_total = 0
        workdays_elapsed = 0
        iter_date = start.date()
        while iter_date <= end.date():
            if iter_date.weekday() < 5:
                workdays_total += 1
                if iter_date <= today:
                    workdays_elapsed += 1
            iter_date += timedelta(days=1)
    
    else:
        raise ValueError(f"Unknown period_type: {period_type}")
    
    return start, end, workdays_total, workdays_elapsed


def _fetch_calendar_events(
    start: datetime,
    end: datetime,
    calendar_id: str,
    account: Optional[str] = None,
) -> list[dict]:
    """Fetch events from Google Calendar."""
    service = get_service(account)
    
    events_result = service.events().list(
        calendarId=calendar_id,
        timeMin=start.isoformat() + "Z" if start.tzinfo is None else start.isoformat(),
        timeMax=end.isoformat() + "Z" if end.tzinfo is None else end.isoformat(),
        singleEvents=True,
        orderBy="startTime",
        maxResults=500,
    ).execute()
    
    return events_result.get("items", [])


def _calculate_metrics(
    entries: list[TimeEntry],
    norm_hours: int,
    billable_target_type: str,
    billable_target_value: float,
    workdays_elapsed: int,
) -> dict:
    """Calculate time tracking metrics from parsed entries."""
    
    # Filter out excluded events
    active_entries = [e for e in entries if not e.is_excluded]
    valid_entries = [e for e in active_entries if e.is_valid]
    error_entries = [e for e in active_entries if e.has_errors]
    
    # Total hours
    total_hours = sum(e.duration_hours for e in active_entries)
    total_valid_hours = sum(e.duration_hours for e in valid_entries)
    total_error_hours = sum(e.duration_hours for e in error_entries)
    
    # Billable breakdown
    billable_hours = sum(e.duration_hours for e in valid_entries if e.is_billable)
    non_billable_hours = sum(e.duration_hours for e in valid_entries if not e.is_billable)
    
    # Billable with full tracking (phase + task)
    billable_full = sum(
        e.duration_hours for e in valid_entries
        if e.is_billable and e.phase_code and e.task_code
    )
    billable_phase_only = sum(
        e.duration_hours for e in valid_entries
        if e.is_billable and e.phase_code and not e.task_code
    )
    billable_project_only = sum(
        e.duration_hours for e in valid_entries
        if e.is_billable and not e.phase_code and not e.task_code
    )
    
    # Calculate targets
    elapsed_norm = workdays_elapsed * 8 if workdays_elapsed > 0 else 8
    
    if billable_target_type == "percent":
        billable_target_hours = norm_hours * (billable_target_value / 100)
        elapsed_billable_target = elapsed_norm * (billable_target_value / 100)
    else:  # days
        billable_target_hours = billable_target_value * 8
        elapsed_billable_target = min(billable_target_hours, elapsed_norm)
    
    # Calculate percentages
    total_pct_of_norm = (total_valid_hours / norm_hours * 100) if norm_hours > 0 else 0
    total_pct_of_elapsed = (total_valid_hours / elapsed_norm * 100) if elapsed_norm > 0 else 0
    
    billable_pct_of_total = (billable_hours / total_valid_hours * 100) if total_valid_hours > 0 else 0
    billable_pct_of_target = (billable_hours / billable_target_hours * 100) if billable_target_hours > 0 else 0
    billable_pct_of_elapsed_target = (billable_hours / elapsed_billable_target * 100) if elapsed_billable_target > 0 else 0
    
    return {
        "total_hours": round(total_hours, 2),
        "total_valid_hours": round(total_valid_hours, 2),
        "total_error_hours": round(total_error_hours, 2),
        "billable_hours": round(billable_hours, 2),
        "non_billable_hours": round(non_billable_hours, 2),
        "billable_full_tracking": round(billable_full, 2),
        "billable_phase_only": round(billable_phase_only, 2),
        "billable_project_only": round(billable_project_only, 2),
        "norm_hours": norm_hours,
        "elapsed_norm_hours": elapsed_norm,
        "billable_target_hours": round(billable_target_hours, 2),
        "total_pct_of_norm": round(total_pct_of_norm, 1),
        "total_pct_of_elapsed": round(total_pct_of_elapsed, 1),
        "billable_pct_of_total": round(billable_pct_of_total, 1),
        "billable_pct_of_target": round(billable_pct_of_target, 1),
        "billable_pct_of_elapsed_target": round(billable_pct_of_elapsed_target, 1),
        "entries_total": len(active_entries),
        "entries_valid": len(valid_entries),
        "entries_error": len(error_entries),
    }


def _group_by_project(entries: list[TimeEntry]) -> dict:
    """Group entries by project with hours breakdown."""
    projects = {}
    
    for entry in entries:
        if entry.is_excluded:
            continue
        
        code = entry.project_code or "UNTRACKED"
        if code not in projects:
            projects[code] = {
                "hours": 0,
                "billable": entry.is_billable if entry.project_code else False,
                "entries": 0,
            }
        
        projects[code]["hours"] += entry.duration_hours
        projects[code]["entries"] += 1
    
    # Round hours
    for code in projects:
        projects[code]["hours"] = round(projects[code]["hours"], 2)
    
    return dict(sorted(projects.items(), key=lambda x: x[1]["hours"], reverse=True))


def _get_error_details(entries: list[TimeEntry]) -> list[dict]:
    """Extract error entries for reporting."""
    errors = []
    for entry in entries:
        if entry.has_errors and not entry.is_excluded:
            errors.append({
                "date": entry.date.strftime("%Y-%m-%d"),
                "hours": entry.duration_hours,
                "summary": entry.raw_summary[:100],
                "errors": entry.errors,
            })
    return errors


async def time_tracking_report(
    period_type: str = "month",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    output_format: str = "summary",
    account: Optional[str] = None,
) -> dict:
    """
    Generate time tracking report.
    
    Args:
        period_type: 'week' (week-to-date), 'month' (month-to-date), or 'custom'
        start_date: Start date for custom period (YYYY-MM-DD)
        end_date: End date for custom period (YYYY-MM-DD)
        output_format: 'summary' (text metrics) or 'excel' (file path)
        account: Google account name (uses default if not specified)
    
    Returns:
        Dict with:
        - period: date range info
        - metrics: hours and percentages
        - by_project: breakdown by project
        - errors: list of unparseable entries
        - file_path: (if output_format='excel') path to generated file
    
    Metrics explained:
        - total_pct_of_norm: Progress toward full month goal
        - total_pct_of_elapsed: Whether on track for elapsed time
        - billable_pct_of_target: Progress toward billable goal
        - billable_pct_of_elapsed_target: On track for billable (should be ~100%)
    """
    ensure_database()
    
    # Get settings
    calendar_id = get_setting("work_calendar") or "primary"
    billable_target_type = get_setting("billable_target_type") or "percent"
    billable_target_value = float(get_setting("billable_target_value") or "75")
    
    # Calculate period
    try:
        start, end, workdays_total, workdays_elapsed = _get_period_bounds(
            period_type, start_date, end_date
        )
    except ValueError as e:
        return {"error": str(e)}
    
    # Get norm hours
    norm_data = get_norm(start.year, start.month)
    if norm_data:
        norm_hours = norm_data["norm_hours"]
    else:
        # Default: workdays × 8
        norm_hours = workdays_total * 8
    
    # Adjust norm for week reports
    if period_type == "week":
        norm_hours = 40  # Standard week
    
    # Fetch events
    try:
        events = _fetch_calendar_events(start, end, calendar_id, account)
    except Exception as e:
        return {"error": f"Failed to fetch calendar events: {str(e)}"}
    
    # Parse events
    entries = parse_events_batch(events)
    
    # Calculate metrics
    metrics = _calculate_metrics(
        entries,
        norm_hours,
        billable_target_type,
        billable_target_value,
        workdays_elapsed,
    )
    
    # Group by project
    by_project = _group_by_project(entries)
    
    # Get errors
    errors = _get_error_details(entries)
    
    result = {
        "period": {
            "type": period_type,
            "start": start.strftime("%Y-%m-%d"),
            "end": end.strftime("%Y-%m-%d"),
            "workdays_total": workdays_total,
            "workdays_elapsed": workdays_elapsed,
        },
        "metrics": metrics,
        "by_project": by_project,
        "errors": errors,
        "settings": {
            "calendar_id": calendar_id,
            "billable_target_type": billable_target_type,
            "billable_target_value": billable_target_value,
        }
    }
    
    # Generate Excel if requested
    if output_format == "excel":
        try:
            file_path = _generate_excel_report(result, entries, period_type)
            result["file_path"] = str(file_path)
        except Exception as e:
            result["excel_error"] = str(e)
    
    return result


def _generate_excel_report(
    report_data: dict,
    entries: list[TimeEntry],
    period_type: str,
) -> Path:
    """
    Generate Excel timesheet report matching original iCal format.
    
    Columns (matching 1C import format):
    A: Date
    B: Fact hours
    C: Project
    D: Project phase
    E: (empty)
    F: Location
    G: Description
    H: Per diems
    I: Title (position)
    J: Comment
    K: Errors
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")
    
    # Get base_location from settings
    base_location = get_setting("base_location") or ""
    
    wb = Workbook()
    
    # Create sheet
    ws = wb.active
    ws.title = f"{period_type}-to-date"
    
    # Headers matching 1C import format (10 columns)
    headers = ["Date", "Fact hours", "Project", "Project phase", "Location", "Description", "Per diems", "Title", "Comment", "Errors"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Data rows
    row = 2
    for entry in entries:
        if entry.is_excluded:
            continue
        
        # Build description: task_code * description if task exists, else just description
        if entry.task_code and entry.description:
            description = f"{entry.task_code} * {entry.description}"
        elif entry.description:
            description = entry.description
        else:
            description = entry.raw_summary[:100] if entry.raw_summary else ""
        
        # A: Date
        ws.cell(row=row, column=1, value=entry.date.strftime("%d.%m.%Y"))
        # B: Fact hours
        ws.cell(row=row, column=2, value=entry.duration_hours)
        # C: Project
        ws.cell(row=row, column=3, value=entry.project_code or "")
        # D: Project phase
        ws.cell(row=row, column=4, value=entry.phase_code or "")
        # E: Location
        ws.cell(row=row, column=5, value=base_location)
        # F: Description
        ws.cell(row=row, column=6, value=description)
        # G: Per diems (empty)
        ws.cell(row=row, column=7, value="")
        # H: Title (position from project)
        ws.cell(row=row, column=8, value=entry.position or "")
        # I: Comment (empty)
        ws.cell(row=row, column=9, value="")
        # J: Errors
        ws.cell(row=row, column=10, value="; ".join(entry.errors) if entry.errors else "")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 80
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 30
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    metrics = report_data["metrics"]
    period = report_data["period"]
    
    summary_data = [
        ["Time Tracking Report", ""],
        ["Period", f"{period['start']} to {period['end']}"],
        ["Type", period_type.upper()],
        ["", ""],
        ["Total Hours", metrics["total_valid_hours"]],
        ["Norm Hours", metrics["norm_hours"]],
        ["Progress %", f"{metrics['total_pct_of_norm']}%"],
        ["On-track %", f"{metrics['total_pct_of_elapsed']}%"],
        ["", ""],
        ["Billable Hours", metrics["billable_hours"]],
        ["Non-billable Hours", metrics["non_billable_hours"]],
        ["Billable Target", metrics["billable_target_hours"]],
        ["Billable Progress %", f"{metrics['billable_pct_of_target']}%"],
        ["Billable On-track %", f"{metrics['billable_pct_of_elapsed_target']}%"],
        ["", ""],
        ["Entries Total", metrics["entries_total"]],
        ["Entries Valid", metrics["entries_valid"]],
        ["Entries with Errors", metrics["entries_error"]],
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        summary_ws.cell(row=row_idx, column=1, value=label)
        summary_ws.cell(row=row_idx, column=2, value=value)
    
    summary_ws.column_dimensions["A"].width = 20
    summary_ws.column_dimensions["B"].width = 25
    
    # Save file to Downloads/Timesheet reports/
    reports_dir = get_reports_dir()
    
    filename = f"{datetime.now().strftime('%Y-%m-%d')}_{period_type}_timesheet.xlsx"
    file_path = reports_dir / filename
    
    wb.save(file_path)
    
    return file_path
