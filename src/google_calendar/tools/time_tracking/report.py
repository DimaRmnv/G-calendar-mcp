"""
Report tool for time tracking.

Combines status (quick summary) and full reports with Excel export.
"""

from typing import Optional
from datetime import datetime, timedelta
from pathlib import Path

from google_calendar.tools.time_tracking.database import (
    ensure_database,
    config_get,
    norm_get,
)
from google_calendar.tools.time_tracking.parser import parse_events_batch, TimeEntry
from google_calendar.api.client import get_service


REPORTS_DIR = Path.home() / "Downloads" / "Timesheet reports"


def _get_reports_dir() -> Path:
    """Get reports directory, creating if needed."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return REPORTS_DIR


def _count_workdays(start_date, end_date) -> int:
    """Count workdays between dates (inclusive)."""
    count = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:
            count += 1
        current += timedelta(days=1)
    return count


def _fetch_events(service, calendar_id: str, start: datetime, end: datetime) -> list[dict]:
    """Fetch events from Google Calendar."""
    events_result = service.events().list(
        calendarId=calendar_id,
        timeMin=start.isoformat() + "Z" if start.tzinfo is None else start.isoformat(),
        timeMax=end.isoformat() + "Z" if end.tzinfo is None else end.isoformat(),
        singleEvents=True,
        orderBy="startTime",
        maxResults=500,
    ).execute()
    return events_result.get("items", [])


def _calculate_summary(
    entries: list[TimeEntry],
    period_type: str,
    start: datetime,
    end: datetime,
    norm_hours: float,
    workdays_elapsed: int,
    billable_target_days: int,
) -> dict:
    """Calculate full summary metrics."""
    
    # Filter entries
    active = [e for e in entries if not e.is_excluded]
    valid = [e for e in active if e.is_valid]
    errors = [e for e in active if e.has_errors]
    
    # Hours calculations
    hours_elapsed = workdays_elapsed * 8
    billable_target_hours = billable_target_days * 8
    
    # Total hours (valid entries only)
    total_hours = sum(e.duration_hours for e in valid)
    
    # Billable breakdown
    billable_entries = [e for e in valid if e.is_billable]
    billable_hours = sum(e.duration_hours for e in billable_entries)
    
    billable_with_phase_and_task = sum(
        e.duration_hours for e in billable_entries 
        if e.phase_code and e.task_code
    )
    billable_with_phase_no_task = sum(
        e.duration_hours for e in billable_entries 
        if e.phase_code and not e.task_code
    )
    billable_without_phase = sum(
        e.duration_hours for e in billable_entries 
        if not e.phase_code
    )
    
    # Non-billable
    non_billable_hours = sum(e.duration_hours for e in valid if not e.is_billable)
    
    # Error hours
    error_hours = sum(e.duration_hours for e in errors)
    total_reported = total_hours + error_hours
    
    # Percentage calculations (safe division)
    def safe_pct(numerator, denominator):
        return round(numerator / denominator * 100, 1) if denominator > 0 else 0.0
    
    # Elapsed target for billable (proportional to elapsed time)
    elapsed_billable_target = hours_elapsed * billable_target_hours / norm_hours if norm_hours > 0 else 0
    
    return {
        "period": {
            "type": period_type,
            "start": start.strftime("%Y-%m-%d"),
            "end": end.strftime("%Y-%m-%d"),
            "norm_hours": norm_hours,
            "workdays_elapsed": workdays_elapsed,
            "hours_elapsed": hours_elapsed,
            "billable_target_days": billable_target_days,
            "billable_target_hours": billable_target_hours,
        },
        "total": {
            "hours": round(total_hours, 2),
            "pct_of_monthly_norm": safe_pct(total_hours, norm_hours),
            "pct_of_elapsed_norm": safe_pct(total_hours, hours_elapsed),
        },
        "billable": {
            "hours": round(billable_hours, 2),
            "pct_of_total_worked": safe_pct(billable_hours, total_hours),
            "pct_of_monthly_target": safe_pct(billable_hours, billable_target_hours),
            "pct_of_elapsed_target": safe_pct(billable_hours, elapsed_billable_target),
            "with_phase_and_task": round(billable_with_phase_and_task, 2),
            "with_phase_no_task": round(billable_with_phase_no_task, 2),
            "without_phase": round(billable_without_phase, 2),
        },
        "non_billable": {
            "hours": round(non_billable_hours, 2),
            "pct_of_total_worked": safe_pct(non_billable_hours, total_hours),
        },
        "errors": {
            "hours": round(error_hours, 2),
            "count": len(errors),
            "pct_of_total_reported": safe_pct(error_hours, total_reported),
        },
    }


def _get_error_records(entries: list[TimeEntry]) -> list[dict]:
    """Extract error records with details."""
    errors = [e for e in entries if e.has_errors and not e.is_excluded]
    
    records = []
    for e in errors:
        # Determine billable status from raw summary or default to False
        billable = False
        if e.project_code:
            billable = e.is_billable
        
        records.append({
            "date": e.date.strftime("%d.%m.%Y"),
            "hours": e.duration_hours,
            "project": e.project_code,
            "phase": e.phase_code,
            "description": e.description or e.raw_summary,
            "billable": billable,
            "error": e.errors[0] if e.errors else "Unknown error",
        })
    
    return records


def _generate_excel(entries: list[TimeEntry], period_type: str) -> tuple[Path, str]:
    """Generate Excel timesheet report. Returns (file_path, file_name)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl required. Install: pip install openpyxl")

    base_location = config_get("base_location") or ""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{period_type}-to-date"

    # Headers (10 columns for 1C import)
    headers = ["Date", "Fact hours", "Project", "Project phase", "Location", 
               "Description", "Per diems", "Title", "Comment", "Errors"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data rows
    row = 2
    for entry in entries:
        if entry.is_excluded:
            continue
        
        # Build description: TASK * Description or just description
        if entry.task_code and entry.description:
            desc = f"{entry.task_code} * {entry.description}"
        else:
            desc = entry.description or entry.raw_summary[:100]
        
        ws.cell(row=row, column=1, value=entry.date.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=2, value=entry.duration_hours)
        ws.cell(row=row, column=3, value=entry.project_code or "")
        ws.cell(row=row, column=4, value=entry.phase_code or "")
        ws.cell(row=row, column=5, value=base_location)
        ws.cell(row=row, column=6, value=desc)
        ws.cell(row=row, column=7, value="")  # Per diems
        ws.cell(row=row, column=8, value=entry.position or "")
        ws.cell(row=row, column=9, value="")  # Comment
        ws.cell(row=row, column=10, value="; ".join(entry.errors) if entry.errors else "")
        row += 1

    # Column widths
    widths = [12, 10, 12, 15, 12, 80, 10, 30, 15, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Save
    reports_dir = _get_reports_dir()
    file_name = f"{datetime.now().strftime('%Y-%m-%d')}_{period_type}_timesheet.xlsx"
    file_path = reports_dir / file_name
    wb.save(file_path)
    
    return file_path, file_name


async def time_tracking_report(
    report_type: str = "status",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    account: Optional[str] = None,
) -> dict:
    """
    Time tracking report and status tool.

    Args:
        report_type: 'status' (quick WTD/MTD), 'week', 'month', or 'custom'
        start_date: Start date for custom (YYYY-MM-DD)
        end_date: End date for custom (YYYY-MM-DD)
        account: Google account name

    Returns:
        For status: week/month summaries with on-track percentages
        For reports: summary + error_records + Excel file path (auto-saved)
    """
    ensure_database()

    now = datetime.now()
    today = now.date()

    # Settings
    calendar_id = config_get("work_calendar") or "primary"
    billable_target_type = config_get("billable_target_type") or "days"
    billable_target_value = float(config_get("billable_target_value") or "15")
    
    # Convert to days if needed
    if billable_target_type == "days":
        billable_target_days = int(billable_target_value)
    else:
        # Assume percent, convert to equivalent days (based on 22 workday month)
        billable_target_days = int(22 * billable_target_value / 100)

    # Get service
    try:
        service = get_service(account)
    except Exception as e:
        return {"error": f"Failed to get calendar service: {str(e)}"}

    # STATUS: Quick WTD/MTD summary (unchanged logic)
    if report_type == "status":
        week_start = now - timedelta(days=now.weekday())
        week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        month_start = datetime(now.year, now.month, 1)
        end = now.replace(hour=23, minute=59, second=59)

        week_workdays = _count_workdays(week_start.date(), today)
        month_workdays = _count_workdays(month_start.date(), today)

        norm_data = norm_get(year=now.year, month=now.month)
        month_norm = norm_data["hours"] if norm_data else month_workdays * 8
        billable_target_hours = billable_target_days * 8

        try:
            events = _fetch_events(service, calendar_id, month_start, end)
        except Exception as e:
            return {"error": f"Failed to fetch events: {str(e)}"}

        entries = parse_events_batch(events)
        active = [e for e in entries if not e.is_excluded]

        week_start_date = week_start.date()
        week_entries = [e for e in active if e.date.date() >= week_start_date]
        month_entries = active

        # Week metrics
        week_valid = [e for e in week_entries if e.is_valid]
        week_total = sum(e.duration_hours for e in week_valid)
        week_billable = sum(e.duration_hours for e in week_valid if e.is_billable)
        week_errors = len([e for e in week_entries if e.has_errors])
        week_elapsed = week_workdays * 8
        week_billable_target = week_elapsed * billable_target_hours / month_norm if month_norm > 0 else 0

        # Month metrics
        month_valid = [e for e in month_entries if e.is_valid]
        month_total = sum(e.duration_hours for e in month_valid)
        month_billable = sum(e.duration_hours for e in month_valid if e.is_billable)
        month_errors = len([e for e in month_entries if e.has_errors])
        month_elapsed = month_workdays * 8
        month_billable_target = month_elapsed * billable_target_hours / month_norm if month_norm > 0 else 0

        week_on_track = (week_total / week_elapsed * 100) if week_elapsed > 0 else 0
        week_billable_on_track = (week_billable / week_billable_target * 100) if week_billable_target > 0 else 0
        month_on_track = (month_total / month_elapsed * 100) if month_elapsed > 0 else 0
        month_billable_on_track = (month_billable / month_billable_target * 100) if month_billable_target > 0 else 0

        status_emoji = "✅" if month_on_track >= 95 else "⚠️" if month_on_track >= 80 else "🔴"
        message = (
            f"{status_emoji} MTD: {round(month_total, 1)}h total, {round(month_billable, 1)}h billable "
            f"({round(month_billable_on_track, 0)}% on-track). "
            f"WTD: {round(week_total, 1)}h total, {round(week_billable, 1)}h billable "
            f"({round(week_billable_on_track, 0)}% on-track)."
        )
        if month_errors > 0:
            message += f" ⚠️ {month_errors} events need attention."

        return {
            "week": {
                "total_hours": round(week_total, 2),
                "billable_hours": round(week_billable, 2),
                "pct_of_elapsed_norm": round(week_on_track, 1),
                "pct_of_elapsed_target": round(week_billable_on_track, 1),
                "workdays": week_workdays,
                "errors": week_errors,
            },
            "month": {
                "total_hours": round(month_total, 2),
                "billable_hours": round(month_billable, 2),
                "norm_hours": month_norm,
                "pct_of_elapsed_norm": round(month_on_track, 1),
                "pct_of_elapsed_target": round(month_billable_on_track, 1),
                "workdays": month_workdays,
                "errors": month_errors,
            },
            "message": message,
        }

    # REPORT: Week/Month/Custom
    if report_type == "week":
        start = now - timedelta(days=now.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now.replace(hour=23, minute=59, second=59)
        norm_data = norm_get(year=now.year, month=now.month)
        norm_hours = norm_data["hours"] if norm_data else 176
    elif report_type == "month":
        start = datetime(now.year, now.month, 1)
        end = now.replace(hour=23, minute=59, second=59)
        norm_data = norm_get(year=now.year, month=now.month)
        norm_hours = norm_data["hours"] if norm_data else _count_workdays(start.date(), today) * 8
    elif report_type == "custom":
        if not start_date or not end_date:
            return {"error": "start_date and end_date required for custom report"}
        start = datetime.fromisoformat(start_date)
        end = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59)
        # For custom, use the month norm of start date
        norm_data = norm_get(year=start.year, month=start.month)
        norm_hours = norm_data["hours"] if norm_data else _count_workdays(start.date(), end.date()) * 8
    else:
        return {"error": f"Unknown report_type: {report_type}. Use: status, week, month, custom"}

    workdays_elapsed = _count_workdays(start.date(), min(today, end.date()))

    try:
        events = _fetch_events(service, calendar_id, start, end)
    except Exception as e:
        return {"error": f"Failed to fetch events: {str(e)}"}

    entries = parse_events_batch(events)
    
    # Build response
    summary = _calculate_summary(
        entries=entries,
        period_type=report_type,
        start=start,
        end=end,
        norm_hours=norm_hours,
        workdays_elapsed=workdays_elapsed,
        billable_target_days=billable_target_days,
    )
    
    error_records = _get_error_records(entries)
    
    # Always generate Excel
    try:
        file_path, file_name = _generate_excel(entries, report_type)
    except Exception as e:
        return {
            "summary": summary,
            "error_records": error_records,
            "excel_error": str(e),
        }

    return {
        "summary": summary,
        "error_records": error_records,
        "file_path": str(file_path),
        "file_name": file_name,
    }
