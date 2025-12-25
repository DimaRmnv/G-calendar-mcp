"""
Contact reporting and analytics module.

Provides reports on contacts, project teams, and communication patterns.
Includes Excel export functionality with download URLs.

Uses PostgreSQL via asyncpg.
"""

import os
import uuid as uuid_module
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

from google_calendar.db.connection import get_db


# Base URL for download links (from env, fallback for local dev)
EXPORT_BASE_URL = os.environ.get("EXPORT_BASE_URL", "http://localhost:8000")


async def contacts_report(
    report_type: str,
    project_id: int = None,
    organization: str = None,
    days_stale: int = 90,
    limit: int = 50
) -> dict:
    """
    Generate contact reports.

    Args:
        report_type: Type of report:
            - 'project_team': Full team roster with contacts for a project
            - 'organization': Contacts grouped by organization
            - 'communication_map': Contact channels summary
            - 'stale_contacts': Contacts not updated in X days
            - 'summary': Overall contacts summary
        project_id: Required for project_team report
        organization: Filter for organization report
        days_stale: Days threshold for stale_contacts (default 90)
        limit: Max results (default 50)

    Returns:
        Dict with report data based on report_type
    """
    if report_type == 'project_team':
        return await _report_project_team(project_id)
    elif report_type == 'organization':
        return await _report_by_organization(organization, limit)
    elif report_type == 'communication_map':
        return await _report_communication_map(limit)
    elif report_type == 'stale_contacts':
        return await _report_stale_contacts(days_stale, limit)
    elif report_type == 'summary':
        return await _report_summary()
    else:
        return {"error": f"Unknown report type: {report_type}"}


async def _report_project_team(project_id: int) -> dict:
    """Full team roster for a project with all contact details."""
    if not project_id:
        return {"error": "project_id is required for project_team report"}

    async with get_db() as conn:
        # Get project info
        project = await conn.fetchrow("""
            SELECT id, code, description, is_billable, is_active
            FROM projects WHERE id = $1
        """, project_id)

        if not project:
            return {"error": f"Project {project_id} not found"}

        project_info = dict(project)

        # Get team members with roles
        members_raw = await conn.fetch("""
            SELECT
                c.id as contact_id,
                c.first_name,
                c.last_name,
                c.organization,
                c.job_title,
                c.country,
                cp.role_code,
                pr.role_name_en as role_name,
                pr.role_category,
                cp.start_date,
                cp.end_date,
                cp.workdays_allocated,
                cp.is_active as assignment_active
            FROM contact_projects cp
            JOIN contacts c ON cp.contact_id = c.id
            JOIN project_roles pr ON cp.role_code = pr.role_code
            WHERE cp.project_id = $1
            ORDER BY pr.role_category, pr.role_name_en, c.last_name
        """, project_id)

        members = []
        for row in members_raw:
            member = dict(row)

            # Get channels for this contact
            channels = await conn.fetch("""
                SELECT channel_type, channel_value, is_primary
                FROM contact_channels
                WHERE contact_id = $1
                ORDER BY is_primary DESC
            """, member['contact_id'])

            member['channels'] = [dict(ch) for ch in channels]

            # Extract primary channels
            for ch in member['channels']:
                if ch['is_primary']:
                    member[f"primary_{ch['channel_type']}"] = ch['channel_value']

            members.append(member)

        # Group by role category
        by_category = {}
        for m in members:
            cat = m.get('role_category', 'other')
            if cat not in by_category:
                by_category[cat] = []
            by_category[cat].append(m)

        return {
            'report_type': 'project_team',
            'project': project_info,
            'team_size': len(members),
            'members': members,
            'by_category': by_category,
            'generated_at': datetime.now().isoformat()
        }


async def _report_by_organization(organization: str = None, limit: int = 50) -> dict:
    """Contacts grouped by organization."""
    async with get_db() as conn:
        if organization:
            # Single organization
            contacts_raw = await conn.fetch("""
                SELECT
                    id, first_name, last_name, organization,
                    organization_type, job_title, country,
                    preferred_channel
                FROM contacts
                WHERE organization = $1 OR organization ILIKE $2
                ORDER BY last_name, first_name
                LIMIT $3
            """, organization, f"%{organization}%", limit)

            contacts = [dict(row) for row in contacts_raw]

            # Get channels for each
            for c in contacts:
                channels = await conn.fetch("""
                    SELECT channel_type, channel_value, is_primary
                    FROM contact_channels WHERE contact_id = $1
                """, c['id'])
                c['channels'] = [dict(ch) for ch in channels]

            return {
                'report_type': 'organization',
                'organization': organization,
                'contact_count': len(contacts),
                'contacts': contacts,
                'generated_at': datetime.now().isoformat()
            }
        else:
            # All organizations summary
            orgs = await conn.fetch("""
                SELECT
                    organization,
                    organization_type,
                    COUNT(*) as contact_count,
                    STRING_AGG(DISTINCT country, ', ') as countries
                FROM contacts
                WHERE organization IS NOT NULL AND organization != ''
                GROUP BY organization, organization_type
                ORDER BY contact_count DESC
                LIMIT $1
            """, limit)

            return {
                'report_type': 'organization_summary',
                'organization_count': len(orgs),
                'organizations': [dict(row) for row in orgs],
                'generated_at': datetime.now().isoformat()
            }


async def _report_communication_map(limit: int = 50) -> dict:
    """Summary of contact channels - who can be reached how."""
    async with get_db() as conn:
        # Channel type distribution
        channel_stats = await conn.fetch("""
            SELECT
                channel_type,
                COUNT(*) as count,
                COUNT(CASE WHEN is_primary = TRUE THEN 1 END) as primary_count
            FROM contact_channels
            GROUP BY channel_type
            ORDER BY count DESC
        """)

        # Contacts with multiple channels
        multi_channel = await conn.fetch("""
            SELECT
                c.id,
                c.first_name || ' ' || c.last_name as name,
                c.organization,
                COUNT(cc.id) as channel_count,
                STRING_AGG(DISTINCT cc.channel_type, ', ') as channel_types
            FROM contacts c
            JOIN contact_channels cc ON c.id = cc.contact_id
            GROUP BY c.id, c.first_name, c.last_name, c.organization
            ORDER BY channel_count DESC
            LIMIT $1
        """, limit)

        # Contacts without channels
        no_channels = await conn.fetch("""
            SELECT
                c.id,
                c.first_name || ' ' || c.last_name as name,
                c.organization
            FROM contacts c
            LEFT JOIN contact_channels cc ON c.id = cc.contact_id
            WHERE cc.id IS NULL
            LIMIT $1
        """, limit)

        # Preferred channel distribution
        preferred_stats = await conn.fetch("""
            SELECT
                preferred_channel,
                COUNT(*) as count
            FROM contacts
            WHERE preferred_channel IS NOT NULL
            GROUP BY preferred_channel
            ORDER BY count DESC
        """)

        return {
            'report_type': 'communication_map',
            'channel_distribution': [dict(row) for row in channel_stats],
            'preferred_channels': [dict(row) for row in preferred_stats],
            'multi_channel_contacts': [dict(row) for row in multi_channel],
            'contacts_without_channels': [dict(row) for row in no_channels],
            'no_channel_count': len(no_channels),
            'generated_at': datetime.now().isoformat()
        }


async def _report_stale_contacts(days_stale: int = 90, limit: int = 50) -> dict:
    """Contacts not updated recently."""
    async with get_db() as conn:
        cutoff = (datetime.now() - timedelta(days=days_stale)).isoformat()

        stale_raw = await conn.fetch("""
            SELECT
                c.id,
                c.first_name || ' ' || c.last_name as name,
                c.organization,
                c.updated_at,
                EXTRACT(DAY FROM (NOW() - c.updated_at)) as days_since_update
            FROM contacts c
            WHERE c.updated_at < $1 OR c.updated_at IS NULL
            ORDER BY c.updated_at ASC NULLS FIRST
            LIMIT $2
        """, cutoff, limit)

        stale = [dict(row) for row in stale_raw]

        # Add channels for each
        for c in stale:
            channels = await conn.fetch("""
                SELECT channel_type, channel_value
                FROM contact_channels WHERE contact_id = $1
                LIMIT 3
            """, c['id'])
            c['channels'] = [dict(ch) for ch in channels]

        return {
            'report_type': 'stale_contacts',
            'days_threshold': days_stale,
            'stale_count': len(stale),
            'contacts': stale,
            'generated_at': datetime.now().isoformat()
        }


async def _report_summary() -> dict:
    """Overall contacts database summary."""
    async with get_db() as conn:
        # Total contacts
        total_contacts = await conn.fetchval("SELECT COUNT(*) FROM contacts")

        # Contacts by organization type
        by_org_type = await conn.fetch("""
            SELECT organization_type, COUNT(*) as count
            FROM contacts
            WHERE organization_type IS NOT NULL
            GROUP BY organization_type
            ORDER BY count DESC
        """)

        # Contacts by country
        by_country = await conn.fetch("""
            SELECT country, COUNT(*) as count
            FROM contacts
            WHERE country IS NOT NULL
            GROUP BY country
            ORDER BY count DESC
            LIMIT 10
        """)

        # Total channels
        total_channels = await conn.fetchval("SELECT COUNT(*) FROM contact_channels")

        # Channel types
        channel_types = await conn.fetch("""
            SELECT channel_type, COUNT(*) as count
            FROM contact_channels
            GROUP BY channel_type
            ORDER BY count DESC
        """)

        # Project assignments
        active_assignments = await conn.fetchval(
            "SELECT COUNT(*) FROM contact_projects WHERE is_active = TRUE"
        )

        # Unique projects with assignments
        projects_with_teams = await conn.fetchval(
            "SELECT COUNT(DISTINCT project_id) FROM contact_projects WHERE is_active = TRUE"
        )

        # Recent additions (last 30 days)
        cutoff = (datetime.now() - timedelta(days=30)).isoformat()
        recent_additions = await conn.fetchval(
            "SELECT COUNT(*) FROM contacts WHERE created_at > $1", cutoff
        )

        return {
            'report_type': 'summary',
            'total_contacts': total_contacts,
            'total_channels': total_channels,
            'active_assignments': active_assignments,
            'projects_with_teams': projects_with_teams,
            'recent_additions_30d': recent_additions,
            'by_organization_type': [dict(row) for row in by_org_type],
            'by_country': [dict(row) for row in by_country],
            'channel_types': [dict(row) for row in channel_types],
            'generated_at': datetime.now().isoformat()
        }


async def export_contacts_excel(
    filter_params: dict = None,
    output_path: str = None
) -> dict:
    """
    Export contacts to Excel with all channels.

    Args:
        filter_params: Optional filters:
            - organization: Filter by organization
            - organization_type: Filter by org type
            - country: Filter by country
            - project_id: Filter by project assignment
        output_path: Custom output path (default: ~/Downloads/contacts_export_YYYYMMDD.xlsx)

    Returns:
        Dict with filepath and export stats
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    filter_params = filter_params or {}

    async with get_db() as conn:
        # Build query with filters
        conditions = []
        params = []
        param_idx = 1

        base_query = """
            SELECT DISTINCT
                c.id,
                c.first_name,
                c.last_name,
                c.organization,
                c.organization_type,
                c.job_title,
                c.country,
                c.preferred_channel,
                c.notes,
                c.created_at,
                c.updated_at
            FROM contacts c
        """

        if filter_params.get('project_id'):
            base_query += " JOIN contact_projects cp ON c.id = cp.contact_id"
            conditions.append(f"cp.project_id = ${param_idx}")
            params.append(filter_params['project_id'])
            param_idx += 1

        if filter_params.get('organization'):
            conditions.append(f"c.organization ILIKE ${param_idx}")
            params.append(f"%{filter_params['organization']}%")
            param_idx += 1

        if filter_params.get('organization_type'):
            conditions.append(f"c.organization_type = ${param_idx}")
            params.append(filter_params['organization_type'])
            param_idx += 1

        if filter_params.get('country'):
            conditions.append(f"c.country = ${param_idx}")
            params.append(filter_params['country'])
            param_idx += 1

        if conditions:
            base_query += " WHERE " + " AND ".join(conditions)

        base_query += " ORDER BY c.last_name, c.first_name"

        contacts_raw = await conn.fetch(base_query, *params)
        contacts = [dict(row) for row in contacts_raw]

        # Get all channels
        for c in contacts:
            channels = await conn.fetch("""
                SELECT channel_type, channel_value, is_primary
                FROM contact_channels WHERE contact_id = $1
            """, c['id'])

            # Flatten channels into contact dict
            for ch in channels:
                ch_type = ch['channel_type']
                if ch['is_primary']:
                    c[f'{ch_type}_primary'] = ch['channel_value']
                else:
                    key = ch_type
                    i = 2
                    while key in c:
                        key = f"{ch_type}_{i}"
                        i += 1
                    c[key] = ch['channel_value']

        # Get project assignments
        for c in contacts:
            projects = await conn.fetch("""
                SELECT p.code, pr.role_name_en
                FROM contact_projects cp
                JOIN projects p ON cp.project_id = p.id
                JOIN project_roles pr ON cp.role_code = pr.role_code
                WHERE cp.contact_id = $1 AND cp.is_active = TRUE
            """, c['id'])
            c['projects'] = ', '.join([f"{p['code']} ({p['role_name_en']})" for p in projects])

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Define columns
    columns = [
        'ID', 'First Name', 'Last Name', 'Organization', 'Org Type',
        'Job Title', 'Country', 'Preferred Channel',
        'Email', 'Phone', 'Telegram', 'Teams', 'WhatsApp',
        'Projects', 'Notes', 'Created', 'Updated'
    ]

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for row_idx, contact in enumerate(contacts, 2):
        ws.cell(row=row_idx, column=1, value=contact.get('id'))
        ws.cell(row=row_idx, column=2, value=contact.get('first_name'))
        ws.cell(row=row_idx, column=3, value=contact.get('last_name'))
        ws.cell(row=row_idx, column=4, value=contact.get('organization'))
        ws.cell(row=row_idx, column=5, value=contact.get('organization_type'))
        ws.cell(row=row_idx, column=6, value=contact.get('job_title'))
        ws.cell(row=row_idx, column=7, value=contact.get('country'))
        ws.cell(row=row_idx, column=8, value=contact.get('preferred_channel'))
        ws.cell(row=row_idx, column=9, value=contact.get('email_primary') or contact.get('email'))
        ws.cell(row=row_idx, column=10, value=contact.get('phone_primary') or contact.get('phone'))
        ws.cell(row=row_idx, column=11, value=contact.get('telegram_username') or contact.get('telegram_chat_id'))
        ws.cell(row=row_idx, column=12, value=contact.get('teams_chat_id'))
        ws.cell(row=row_idx, column=13, value=contact.get('whatsapp'))
        ws.cell(row=row_idx, column=14, value=contact.get('projects'))
        ws.cell(row=row_idx, column=15, value=contact.get('notes'))
        ws.cell(row=row_idx, column=16, value=str(contact.get('created_at') or ''))
        ws.cell(row=row_idx, column=17, value=str(contact.get('updated_at') or ''))

        # Apply borders
        for col in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Auto-adjust column widths
    for col in range(1, len(columns) + 1):
        max_length = len(columns[col-1])
        for row in range(2, len(contacts) + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, min(len(str(cell_value)), 50))
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save to disk and create download link
    file_uuid = uuid_module.uuid4().hex
    filename = f"Contacts_Export_{datetime.now().strftime('%d%b%Y')}.xlsx"
    file_path = Path("/data/reports") / f"{file_uuid}.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(file_path)

    # Record in database (TTL = 1 hour)
    expires_at = datetime.now(timezone.utc) + timedelta(hours=1)
    async with get_db() as conn:
        await conn.execute(
            """
            INSERT INTO export_files (uuid, filename, file_path, expires_at)
            VALUES ($1, $2, $3, $4)
            """,
            file_uuid, filename, str(file_path), expires_at
        )

    download_url = f"{EXPORT_BASE_URL}/export/{file_uuid}"

    return {
        'contact_count': len(contacts),
        'filters_applied': filter_params,
        'download_url': download_url,
        'expires_in': '1 hour',
        'filename': filename
    }


async def export_project_team_excel(
    project_id: int,
    output_path: str = None
) -> dict:
    """
    Export project team roster to Excel.

    Args:
        project_id: Project ID
        output_path: Custom output path

    Returns:
        Dict with filepath and export stats
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    # Get team data
    team_data = await _report_project_team(project_id)
    if 'error' in team_data:
        return team_data

    project = team_data['project']
    members = team_data['members']

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Team - {project['code']}"

    # Header styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.cell(row=1, column=1, value=f"Project Team: {project['code']} - {project.get('description', '')}").font = title_font
    ws.merge_cells('A1:H1')

    # Column headers
    columns = [
        'Name', 'Role', 'Category', 'Organization',
        'Email', 'Phone', 'Telegram', 'Teams'
    ]

    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Write members
    for row_idx, member in enumerate(members, 4):
        name = f"{member.get('first_name', '')} {member.get('last_name', '')}".strip()
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=member.get('role_name'))
        ws.cell(row=row_idx, column=3, value=member.get('role_category'))
        ws.cell(row=row_idx, column=4, value=member.get('organization'))

        # Extract channels
        channels = {ch['channel_type']: ch['channel_value'] for ch in member.get('channels', [])}
        ws.cell(row=row_idx, column=5, value=channels.get('email'))
        ws.cell(row=row_idx, column=6, value=channels.get('phone'))
        ws.cell(row=row_idx, column=7, value=channels.get('telegram_username') or channels.get('telegram_chat_id'))
        ws.cell(row=row_idx, column=8, value=channels.get('teams_chat_id'))

        for col in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Auto-adjust column widths
    for col in range(1, len(columns) + 1):
        max_length = len(columns[col-1])
        for row in range(4, len(members) + 4):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, min(len(str(cell_value)), 40))
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    # Freeze header
    ws.freeze_panes = 'A4'

    # Save to disk and create download link
    file_uuid = uuid_module.uuid4().hex
    filename = f"Team_{project['code']}_{datetime.now().strftime('%d%b%Y')}.xlsx"
    file_path = Path("/data/reports") / f"{file_uuid}.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(file_path)

    # Record in database (TTL = 1 hour)
    expires_at = datetime.now(timezone.utc) + timedelta(hours=1)
    async with get_db() as conn:
        await conn.execute(
            """
            INSERT INTO export_files (uuid, filename, file_path, expires_at)
            VALUES ($1, $2, $3, $4)
            """,
            file_uuid, filename, str(file_path), expires_at
        )

    download_url = f"{EXPORT_BASE_URL}/export/{file_uuid}"

    return {
        'project': project,
        'member_count': len(members),
        'download_url': download_url,
        'expires_in': '1 hour',
        'filename': filename
    }
